import os
import csv, io
import zipfile
import tempfile
import re
from io import BytesIO
from functools import wraps
from datetime import datetime, timedelta, timezone

from sqlalchemy.sql import func
from flask_login import login_required, current_user
from sqlalchemy.exc import SQLAlchemyError

from flask import (
    Flask, session, render_template, redirect, url_for,
    request, flash, send_file, jsonify, abort
)
from flask_migrate import Migrate
from flask_sqlalchemy import SQLAlchemy
from flask_login import (
    LoginManager, login_user, login_required,
    logout_user, current_user
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from flask_wtf import FlaskForm
from flask_wtf.csrf import CSRFProtect

from docxtpl import DocxTemplate
from openai import OpenAI
from dotenv import load_dotenv
load_dotenv()

import csv, io, tempfile
from datetime import datetime
from werkzeug.utils import secure_filename

# If you have openpyxl installed, we can read .xlsx; otherwise we’ll allow CSV only.
try:
    import openpyxl  # noqa: F401
    XLSX_ENABLED = True
except Exception:
    XLSX_ENABLED = False

ALLOWED_EXTS = {"csv", "xlsx"} if XLSX_ENABLED else {"csv"}

# Map *your* Employee model fields here.
# Keep this list in sync with your actual columns.
# Map Employee model fields you actually want to import
# Matches your export columns and templates
EMPLOYEE_FIELDS = [
    "first_name", "last_name", "email", "job_title", "line_manager",
    "service", "team_id", "start_date"
]

# ---- Curated concern tags (lightweight, can move to DB later) ----
CURATED_TAGS = {
    "Timekeeping": ["lateness", "missed clock-in", "extended breaks", "early finish", "timekeeping policy"],
    "Attendance": ["unauthorised absence", "short notice absence", "patterns of absence", "fit note", "return to work"],
    "Quality of Work": ["accuracy", "attention to detail", "rework", "documentation", "SOP adherence"],
    "Productivity": ["missed deadlines", "slow throughput", "low output", "prioritisation", "time management"],
    "Conduct": ["inappropriate language", "unprofessional behaviour", "non-compliance", "policy breach", "conflict"],
    "Communication": ["tone", "late replies", "stakeholder updates", "handover quality", "listening"],
    "Teamwork/Collaboration": ["handover gaps", "knowledge sharing", "supporting peers", "collaboration tools"],
    "Compliance/Process": ["data entry errors", "checklist missed", "audit finding", "process deviation"],
    "Customer Service": ["response times", "complaint handling", "service standards", "follow-up"],
    "Health & Safety": ["PPE", "risk assessment", "manual handling", "incident reporting"]
}

def _merge_curated_and_recent(category: str, recent_tags: list[str], cap: int = 30):
    """Merge curated list for category with recent DB tags, de-duplicated, curated first."""
    out, seen = [], set()
    cat_list = CURATED_TAGS.get(category, [])
    for t in cat_list:
        if t.lower() not in seen:
            out.append(t); seen.add(t.lower())
    # Add recent (keep original casing where possible)
    for t in recent_tags:
        if not t: continue
        key = t.lower().strip()
        if key and key not in seen:
            out.append(t.strip()); seen.add(key)
        if len(out) >= cap: break
    return out

# Minimal required fields — adjust if you need more
REQUIRED_FIELDS = ["first_name", "last_name"]

def _read_csv_bytes(file_bytes: bytes):
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = [dict(r) for r in reader]
    headers = reader.fieldnames or []
    return headers, rows

def _read_xlsx_bytes(file_bytes: bytes):
    # Requires openpyxl; already guarded by XLSX_ENABLED
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes))
    sheet = workbook.active
    headers = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append({headers[i]: (row[i] if i < len(headers) else None) for i in range(len(headers))})
    return headers, rows

def _normalize_header(h):
    # Loose normalisation to help mapping (e.g., "First Name" -> "first_name")
    return (h or "").strip().lower().replace(" ", "_")

def _try_parse_date(value):
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except Exception:
            continue
    return None

def _max_wizard_step(data: dict) -> int:
    """
    Gate progression. Only unlock Step 6 (Review) when Action Items exist.
    """
    s = 1
    if data.get('employee_id'): s = 2
    if all(data.get(k) for k in ('concerns', 'concern_category', 'severity', 'frequency')): s = 3
    if all(data.get(k) for k in ('start_date', 'review_date')): s = 4
    if all(data.get(k) for k in ('capability_meeting_date', 'capability_meeting_time', 'capability_meeting_venue')): s = 5
    # Only allow review (6) if at least one action item is present
    items = data.get('action_plan_items') or []
    if s == 5 and isinstance(items, list) and any((x or '').strip() for x in items):
        s = 6
    return s

# --- Models & Forms ---
from models import (
    db, User, Employee, PIPRecord, PIPActionItem, TimelineEvent,
    ProbationRecord, ProbationReview, ProbationPlan, DraftPIP, DraftProbation
)
from forms import (
    PIPForm, EmployeeForm, LoginForm, ProbationRecordForm,
    ProbationReviewForm, ProbationPlanForm, UserForm
)

# Initialize OpenAI v1 client (will pick up OPENAI_API_KEY env var)
client = OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))

# Initialize Flask
app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-only-secret')
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'pip_crm.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
csrf = CSRFProtect(app)

db.init_app(app)
migrate = Migrate(app, db)

# ----- Context Processor -----
@app.context_processor
def inject_module():
    return dict(active_module=session.get('active_module'))

# Initialize Login
login_manager = LoginManager()
login_manager.login_view = 'login'
login_manager.init_app(app)

# ----- User Loader -----
@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# ----- Superuser decorator -----
def superuser_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_superuser():
            abort(403)  # Forbidden
        return f(*args, **kwargs)
    return decorated_function

# ----- Set active module in session -----
@app.before_request
def set_active_module():
    path = request.path.lower()
    if path.startswith('/pip/') or \
       path.startswith('/employee/') or \
       path in ['/dashboard', '/pip_list', '/employee/list', '/employee/add', '/pip/select-employee']:
        session['active_module'] = 'PIP'
    elif path.startswith('/probation/'):
        session['active_module'] = 'Probation'
    elif path == '/':
        session.pop('active_module', None)

# ----- Routes -----
@app.route('/')
@login_required
def home():
    return render_template('select_module.html', hide_sidebar=True, layout='fullscreen')

@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if request.method == 'POST' and form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and check_password_hash(user.password_hash, form.password.data):
            login_user(user, remember=form.remember.data)
            return redirect(url_for('home'))
        flash('Invalid username or password', 'danger')
    return render_template('login.html', form=form, hide_sidebar=True, current_year=datetime.now().year)

# Get active (not dismissed) draft for current user
def get_active_draft_for_user(user_id):
    return DraftPIP.query.filter_by(user_id=user_id, is_dismissed=False)\
                         .order_by(DraftPIP.updated_at.desc()).first()

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if not current_user.is_superuser():
        flash('You do not have permission to access the admin dashboard.', 'danger')
        return redirect(url_for('home'))
    return render_template('admin_dashboard.html')

@app.route('/taxonomy/predefined_tags', methods=['GET'])
@login_required
def taxonomy_predefined_tags():
    cat = (request.args.get('category') or '').strip()
    tags = CURATED_TAGS.get(cat, [])
    return jsonify({"category": cat, "tags": tags})

# ----- Export Data -----
@app.route('/admin/export')
@login_required
@superuser_required
def export_data():
    zip_buffer = BytesIO()

    with tempfile.TemporaryDirectory() as tmpdir:
        def write_csv(filename, fieldnames, rows):
            filepath = os.path.join(tmpdir, filename)
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(rows)
            export_zip.write(filepath, arcname=filename)

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as export_zip:
            # Employees
            employees = Employee.query.all()
            write_csv(
                'employees.csv',
                ['id', 'first_name', 'last_name', 'job_title', 'line_manager', 'service', 'start_date', 'team_id', 'email'],
                [{
                    'id': e.id,
                    'first_name': getattr(e, 'first_name', ''),
                    'last_name': getattr(e, 'last_name', ''),
                    'job_title': e.job_title,
                    'line_manager': e.line_manager,
                    'service': e.service,
                    'start_date': e.start_date.strftime('%Y-%m-%d') if e.start_date else '',
                    'team_id': e.team_id,
                    'email': e.email
                } for e in employees]
            )

            # PIP Records
            pips = PIPRecord.query.all()
            write_csv(
                'pip_records.csv',
                ['id', 'employee_id', 'concerns', 'concern_category', 'severity', 'frequency', 'tags',
                 'start_date', 'review_date', 'status', 'created_by'],
                [{
                    'id': p.id,
                    'employee_id': p.employee_id,
                    'concerns': p.concerns,
                    'concern_category': getattr(p, 'concern_category', ''),
                    'severity': getattr(p, 'severity', ''),
                    'frequency': getattr(p, 'frequency', ''),
                    'tags': getattr(p, 'tags', ''),
                    'start_date': p.start_date.strftime('%Y-%m-%d') if p.start_date else '',
                    'review_date': p.review_date.strftime('%Y-%m-%d') if p.review_date else '',
                    'status': p.status,
                    'created_by': getattr(p, 'created_by', '')
                } for p in pips]
            )

            # Timeline Events
            events = TimelineEvent.query.all()
            write_csv(
                'timeline_events.csv',
                ['id', 'pip_record_id', 'employee_id', 'event_type', 'notes', 'updated_by', 'timestamp'],
                [{
                    'id': t.id,
                    'pip_record_id': t.pip_record_id,
                    'employee_id': t.pip_record.employee_id if t.pip_record else '',
                    'event_type': getattr(t, 'event_type', ''),
                    'notes': getattr(t, 'notes', ''),
                    'updated_by': getattr(t, 'updated_by', ''),
                    'timestamp': t.timestamp.strftime('%Y-%m-%d %H:%M:%S') if t.timestamp else ''
                } for t in events]
            )

            # Users
            users = User.query.all()
            write_csv(
                'users.csv',
                ['id', 'username', 'email', 'admin_level', 'team_id'],
                [{
                    'id': u.id,
                    'username': u.username,
                    'email': u.email,
                    'admin_level': u.admin_level,
                    'team_id': u.team_id
                } for u in users]
            )

            # Probation Records
            probations = ProbationRecord.query.all()
            write_csv(
                'probation_records.csv',
                ['id', 'employee_id', 'status', 'start_date', 'expected_end_date', 'notes'],
                [{
                    'id': p.id,
                    'employee_id': p.employee_id,
                    'status': p.status,
                    'start_date': p.start_date.strftime('%Y-%m-%d') if p.start_date else '',
                    'expected_end_date': p.expected_end_date.strftime('%Y-%m-%d') if p.expected_end_date else '',
                    'notes': p.notes
                } for p in probations]
            )

            # Probation Reviews
            reviews = ProbationReview.query.all()
            write_csv(
                'probation_reviews.csv',
                ['id', 'probation_id', 'review_date', 'reviewer', 'summary', 'concerns_flag'],
                [{
                    'id': r.id,
                    'probation_id': r.probation_id,
                    'review_date': r.review_date.strftime('%Y-%m-%d') if r.review_date else '',
                    'reviewer': r.reviewer,
                    'summary': r.summary,
                    'concerns_flag': r.concerns_flag
                } for r in reviews]
            )

            # Probation Plans (objectives/outcome/deadline)
            plans = ProbationPlan.query.all()
            write_csv(
                'probation_plans.csv',
                ['id', 'probation_id', 'objectives', 'outcome', 'deadline'],
                [{
                    'id': p.id,
                    'probation_id': p.probation_id,
                    'objectives': getattr(p, 'objectives', ''),
                    'outcome': getattr(p, 'outcome', ''),
                    'deadline': p.deadline.strftime('%Y-%m-%d') if getattr(p, 'deadline', None) else ''
                } for p in plans]
            )

    zip_buffer.seek(0)
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M')
    return send_file(zip_buffer, mimetype='application/zip', as_attachment=True, download_name=f'export_{timestamp}.zip')

# ----- User Management -----
@app.route('/admin/users/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    if not current_user.is_superuser():
        flash('Access denied.', 'danger')
        return redirect(url_for('dashboard'))

    user = User.query.get_or_404(user_id)
    form = UserForm(obj=user)

    if form.validate_on_submit():
        user.username = form.username.data
        user.email = form.email.data
        user.admin_level = form.admin_level.data
        user.team_id = form.team_id.data
        db.session.commit()
        flash('User updated successfully.', 'success')
        return redirect(url_for('manage_users'))

    return render_template('edit_user.html', form=form, user=user)

@app.route('/admin/users/create', methods=['GET', 'POST'])
@login_required
def create_user():
    if not current_user.is_superuser():
        flash("Access denied: Superuser only.", "danger")
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        admin_level = int(request.form.get('admin_level', 0))
        team_id = request.form.get('team_id') or None

        if not username or not email or not password:
            flash("All fields except team ID are required.", "danger")
            return redirect(request.url)

        if User.query.filter_by(username=username).first():
            flash("Username already exists.", "danger")
            return redirect(request.url)

        if User.query.filter_by(email=email).first():
            flash("Email already in use.", "danger")
            return redirect(request.url)

        hashed_pw = generate_password_hash(password)
        new_user = User(username=username, email=email, password_hash=hashed_pw, admin_level=admin_level, team_id=team_id)
        db.session.add(new_user)
        db.session.commit()
        flash("User created successfully.", "success")
        return redirect(url_for('manage_users'))

    return render_template('admin_create_user.html')

@app.route('/admin/users/delete/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    if not current_user.is_superuser():
        flash("Access denied: Superuser only.", "danger")
        return redirect(url_for('dashboard'))

    user = User.query.get_or_404(user_id)

    if user.id == current_user.id:
        flash("You cannot delete your own account while logged in.", "warning")
        return redirect(url_for('manage_users'))

    db.session.delete(user)
    db.session.commit()
    flash("User deleted successfully.", "success")
    return redirect(url_for('manage_users'))

# ----- Employee and PIP management -----
@app.route('/employee/<int:employee_id>')
@login_required
def employee_detail(employee_id):
    employee = Employee.query.options(db.joinedload(Employee.pips)).get_or_404(employee_id)
    if current_user.admin_level == 0 and employee.team_id != current_user.team_id:
        flash('Access denied')
        return redirect(url_for('dashboard'))
    return render_template('employee_detail.html', employee=employee)

@app.route('/pip/<int:id>')
@login_required
def pip_detail(id):
    pip = PIPRecord.query.get_or_404(id)
    employee = pip.employee
    return render_template('pip_detail.html', pip=pip, employee=employee)

@app.route('/pip/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_pip(id):
    pip = PIPRecord.query.get_or_404(id)
    employee = pip.employee

    if request.method == 'POST':
        form = PIPForm()
        form.process(request.form)
    else:
        form = PIPForm(obj=pip)
        for _ in range(len(pip.action_items) - len(form.actions.entries)):
            form.actions.append_entry()
        for idx, ai in enumerate(pip.action_items):
            form.actions.entries[idx].form.description.data = ai.description
            form.actions.entries[idx].form.status.data = ai.status

    advice_text = None

    # AI Advice branch
    if request.method == 'POST' and 'generate_advice' in request.form:
        prompt = (
            f"You are a performance coach.\n"
            f"Employee: {employee.first_name} {employee.last_name}\n"
            f"Job Title: {employee.job_title}\n"
            f"Concerns: {form.concerns.data or '[none]'}\n"
            "Action Items:\n"
        )
        for ai_field in form.actions.entries:
            desc = ai_field.form.description.data or '[no description]'
            stat = ai_field.form.status.data or '[no status]'
            prompt += f"- {desc} [{stat}]\n"
        prompt += f"Meeting Notes: {form.meeting_notes.data or '[none]'}\n"
        prompt += "Provide 3 bulleted actionable tips for the manager to support this employee."

        resp = client.chat.completions.create(
            model="gpt-4",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.7,
        )
        advice_text = resp.choices[0].message.content.strip()

        return render_template('edit_pip.html', form=form, pip=pip, employee=employee, advice_text=advice_text)

    # Save branch
    if form.validate_on_submit():
        pip.concerns = form.concerns.data
        pip.start_date = form.start_date.data
        pip.review_date = form.review_date.data
        pip.status = form.status.data
        pip.meeting_notes = form.meeting_notes.data
        pip.capability_meeting_date = form.capability_meeting_date.data
        pip.capability_meeting_time = form.capability_meeting_time.data
        pip.capability_meeting_venue = form.capability_meeting_venue.data

        pip.action_items.clear()
        for ai_field in form.actions.entries:
            pip.action_items.append(
                PIPActionItem(
                    description=ai_field.form.description.data,
                    status=ai_field.form.status.data
                )
            )

        db.session.commit()
        flash('PIP updated successfully.', 'success')
        return redirect(url_for('pip_detail', id=pip.id))

    # Final render
    return render_template('edit_pip.html', form=form, pip=pip, employee=employee, advice_text=advice_text)

@app.route('/pip/<int:id>/generate/advice', methods=['POST'])
@login_required
def generate_ai_advice(id):
    pip = PIPRecord.query.get_or_404(id)
    employee = pip.employee

    prompt = (
        "You are an experienced HR Line Manager based in the UK. "
        "Using the information below, provide three clear and practical suggestions for how a line manager "
        "can support the employee's performance improvement. Your advice should reflect HR best practice in the UK.\n\n"
        f"Employee Name: {employee.first_name} {employee.last_name}\n"
        f"Job Title: {employee.job_title}\n"
        f"Concerns: {pip.concerns or '[none]'}\n"
        "Action Items:\n"
    )
    for item in pip.action_items:
        prompt += f"- {item.description or '[no description]'} [{item.status or '[no status]'}]\n"
    prompt += f"Meeting Notes: {pip.meeting_notes or '[none]'}\n\n"
    prompt += "Provide your advice as a bullet-pointed list."

    resp = client.chat.completions.create(
        model="gpt-4",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.7,
    )
    pip.ai_advice = resp.choices[0].message.content.strip()
    pip.ai_advice_generated_at = datetime.now(timezone.utc)

    event = TimelineEvent(
        pip_record_id=pip.id,
        event_type="AI Advice Generated",
        notes="Advice generated using OpenAI",
        updated_by=current_user.username
    )
    db.session.add(event)

    db.session.commit()
    flash('AI advice generated.', 'success')
    return redirect(url_for('pip_detail', id=pip.id))

@app.route('/pip/create/<int:employee_id>', methods=['GET', 'POST'])
@login_required
def create_pip(employee_id):
    employee = Employee.query.get_or_404(employee_id)
    if current_user.admin_level == 0 and employee.team_id != current_user.team_id:
        flash('Access denied.')
        return redirect(url_for('dashboard'))

    form = PIPForm()

    # Append empty action entry on GET to prevent Jinja error
    if request.method == 'GET' and len(form.actions.entries) == 0:
        form.actions.append_entry()

    # Recalculate min_entries for dynamic action field JS handling
    if request.method == 'POST':
        action_fields = [k for k in request.form if 'actions-' in k and '-description' in k]
        form.actions.min_entries = len(set(k.split('-')[1] for k in action_fields))

    if form.validate_on_submit():
        pip = PIPRecord(
            employee_id=employee.id,
            concerns=form.concerns.data,
            start_date=form.start_date.data,
            review_date=form.review_date.data,
            meeting_notes=form.meeting_notes.data
        )
        db.session.add(pip)
        db.session.flush()

        for action_form in form.actions.entries:
            item = PIPActionItem(
                pip_record_id=pip.id,
                description=action_form.form.description.data,
                status=action_form.form.status.data
            )
            db.session.add(item)

        db.session.commit()
        flash('New PIP created.')
        return redirect(url_for('employee_detail', employee_id=employee.id))

    return render_template('create_pip.html', form=form, employee=employee)

@app.route('/employee/edit/<int:employee_id>', methods=['GET', 'POST'])
@login_required
def edit_employee(employee_id):
    employee = Employee.query.get_or_404(employee_id)

    if current_user.admin_level == 0 and employee.team_id != current_user.team_id:
        flash('Access denied.')
        return redirect(url_for('dashboard'))

    form = EmployeeForm(obj=employee)

    if form.validate_on_submit():
        employee.first_name = form.first_name.data
        employee.last_name = form.last_name.data
        employee.job_title = form.job_title.data
        employee.line_manager = form.line_manager.data
        employee.service = form.service.data
        employee.start_date = form.start_date.data
        employee.team_id = form.team_id.data
        employee.email = form.email.data

        db.session.commit()
        flash('Employee details updated.', 'success')
        return redirect(url_for('employee_detail', employee_id=employee.id))

    return render_template('edit_employee.html', form=form, employee=employee)

@app.route('/pip/list')
@login_required
def pip_list():
    pips = PIPRecord.query.join(Employee).all()
    return render_template('pip_list.html', pips=pips)

# -------- Wizard entry ----------
@app.route("/probation/create-wizard", methods=["GET"])
@login_required
def probation_create_wizard():
    # Most recent active draft for this user
    draft = (DraftProbation.query
             .filter_by(user_id=current_user.id, is_dismissed=False)
             .order_by(DraftProbation.updated_at.desc())
             .first())

    # Optional preselect (?employee_id=123)
    preselect_employee_id = request.args.get("employee_id", type=int)

    # Employees for Step 1 selector
    employees = Employee.query.order_by(Employee.last_name.asc(), Employee.first_name.asc()).all()

    return render_template(
        "create_probation_wizard.html",
        draft=draft,
        employees=employees,
        preselect_employee_id=preselect_employee_id,
        module="Probation"
    )

def _parse_iso_date(s):
    try:
        return datetime.strptime((s or "").strip(), "%Y-%m-%d").date()
    except Exception:
        return None

def _flatten_draft_payload(draft):
    flat = {}
    for _, v in (draft.payload or {}).items():
        if isinstance(v, dict):
            flat.update(v)
    return flat





# -------- Save draft (AJAX) ----------
@app.route("/probation/save_draft", methods=["POST"])
@login_required
@csrf.exempt 
def probation_save_draft():
    data = request.get_json() or {}
    step = int(data.get("step", 1))
    payload_delta = data.get("payload", {})  # only the current step’s data
    draft_id = data.get("draft_id")
    employee_id = data.get("employee_id")

    # Find or create draft (limit to one active per user)
    draft = None
    if draft_id:
        draft = DraftProbation.query.filter_by(id=draft_id, user_id=current_user.id).first()
    if not draft:
        draft = (DraftProbation.query
                 .filter_by(user_id=current_user.id, is_dismissed=False)
                 .order_by(DraftProbation.updated_at.desc())
                 .first())
    if not draft:
        draft = DraftProbation(user_id=current_user.id)

    # merge payload
    merged = draft.payload or {}
    merged[str(step)] = payload_delta
    draft.payload = merged
    draft.step = max(draft.step or 1, step)
    if employee_id:
        draft.employee_id = employee_id

    try:
        db.session.add(draft)
        db.session.commit()
        return jsonify({"ok": True, "draft_id": draft.id,
                        "saved_at": datetime.utcnow().isoformat() + "Z"})
    except SQLAlchemyError as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": "DB_SAVE_FAILED"}), 500

# -------- Validate step (AJAX) ----------
@app.route("/probation/validate_wizard_step", methods=["POST"])
@login_required
@csrf.exempt  
def probation_validate_wizard_step():
    data = request.get_json() or {}
    step = int(data.get("step", 1))
    payload = data.get("payload", {})

    errors = {}

    if step == 1:
        required = ["employee_id", "probation_start", "probation_end"]
        for f in required:
            if not payload.get(f):
                errors[f] = "Required"
    elif step == 2:
        if not payload.get("concern_category") and not payload.get("tags"):
            errors["concern_category"] = "Pick a category or add tags"
    elif step == 3:
        if not payload.get("expected_standards"):
            errors["expected_standards"] = "Add minimum standards"
    elif step == 4:
        # At least one actionable item
        items = payload.get("action_items") or []
        if len(items) == 0:
            errors["action_items"] = "Add at least one action"
    elif step == 5:
        dates = payload.get("review_dates") or []
        if len(dates) == 0:
            errors["review_dates"] = "Add one or more review dates"

    return jsonify({"ok": len(errors) == 0, "errors": errors})

# Helper to finalize (call inside your submit endpoint or within the wizard route)
def finalize_probation_from_draft(draft: 'DraftProbation'):
    """
    Create Probation records from a DraftProbation payload, then dismiss the draft.
    Returns the created ProbationRecord.
    """
    data = _flatten_draft_payload(draft)

    employee_id = draft.employee_id or data.get("employee_id")
    if not employee_id:
        raise ValueError("Missing employee_id in draft")

    # Map wizard fields → ProbationRecord columns
    start_date = _parse_iso_date(data.get("probation_start"))
    end_date   = _parse_iso_date(data.get("probation_end"))

    # Roll some useful context into notes
    notes_parts = []
    if data.get("summary_of_concerns"):
        notes_parts.append(f"Concerns: {data.get('summary_of_concerns')}")
    if data.get("expected_standards"):
        notes_parts.append(f"Standards: {data.get('expected_standards')}")
    if data.get("evidence_notes"):
        notes_parts.append(f"Evidence: {data.get('evidence_notes')}")
    notes = "\n\n".join(notes_parts) if notes_parts else None

    probation = ProbationRecord(
        employee_id=int(employee_id),
        start_date=start_date,
        expected_end_date=end_date,
        notes=notes,
        status="Open"
    )
    db.session.add(probation)
    db.session.flush()  # get probation.id

    # Actions → ProbationPlan(objectives, outcome, deadline)
    for item in (data.get("action_items") or []):
        if not isinstance(item, dict):
            continue
        objectives = (item.get("action") or "").strip()
        if not objectives:
            continue
        deadline = _parse_iso_date(item.get("due_date"))
        plan = ProbationPlan(
            probation_id=probation.id,
            objectives=objectives,
            outcome=(item.get("metric") or "").strip() or None,
            deadline=deadline
        )
        db.session.add(plan)

    # Reviews
    reviewer = (data.get("reviewer") or getattr(current_user, "username", "") or "").strip()
    for d in (data.get("review_dates") or []):
        rd = _parse_iso_date(d)
        if not rd:
            continue
        review = ProbationReview(
            probation_id=probation.id,
            review_date=rd,
            reviewer=reviewer,
            summary=""
        )
        db.session.add(review)

    # Timeline event
    evt = TimelineEvent(
        pip_record_id=None,
        event_type="Probation Created",
        notes=f"Created from wizard by {current_user.username}",
        updated_by=current_user.username
    )
    db.session.add(evt)

    # Dismiss draft
    draft.is_dismissed = True
    draft.updated_at = datetime.now(timezone.utc)

    db.session.commit()
    return probation

# -------- Finalize draft (AJAX) ----------
@app.route("/probation/finalize_from_draft", methods=["POST"])
@login_required
def probation_finalize_from_draft():
    payload = request.get_json(force=True, silent=True) or {}
    draft_id = payload.get("draft_id")
    if not draft_id:
        return jsonify({"ok": False, "error": "NO_DRAFT_ID"}), 400

    draft = (DraftProbation.query
             .filter_by(id=draft_id, user_id=current_user.id, is_dismissed=False)
             .first())
    if not draft:
        return jsonify({"ok": False, "error": "DRAFT_NOT_FOUND"}), 404

    try:
        probation = finalize_probation_from_draft(draft)
        return jsonify({
            "ok": True,
            "probation_id": probation.id,
            "redirect": url_for("view_probation", id=probation.id)
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 400


from flask_wtf.csrf import generate_csrf

@app.context_processor
def inject_csrf_token():
    return dict(csrf_token=generate_csrf)




@app.route('/dashboard')
@login_required
def dashboard():
    total_employees = Employee.query.count()
    active_pips = PIPRecord.query.filter_by(status='Open').count()
    completed_pips = PIPRecord.query.filter_by(status='Completed').count()

    today = datetime.now(timezone.utc).date()

    overdue_reviews = PIPRecord.query.filter(
        PIPRecord.review_date < today,
        PIPRecord.status == 'Open'
    ).count()

    recent_activity = TimelineEvent.query.order_by(TimelineEvent.timestamp.desc()).limit(10).all()
    upcoming_deadline = today + timedelta(days=7)

    if current_user.admin_level == 0:
        upcoming_pips = PIPRecord.query.join(Employee).filter(
            Employee.team_id == current_user.team_id,
            PIPRecord.status == 'Open',
            PIPRecord.review_date >= today,
            PIPRecord.review_date <= upcoming_deadline
        ).order_by(PIPRecord.review_date).all()
    else:
        upcoming_pips = PIPRecord.query.filter(
            PIPRecord.status == 'Open',
            PIPRecord.review_date >= today,
            PIPRecord.review_date <= upcoming_deadline
        ).order_by(PIPRecord.review_date).all()

    draft = get_active_draft_for_user(current_user.id)

    return render_template(
        'dashboard.html',
        total_employees=total_employees,
        active_pips=active_pips,
        completed_pips=completed_pips,
        overdue_reviews=overdue_reviews,
        recent_activity=recent_activity,
        upcoming_pips=upcoming_pips,
        draft=draft
    )

@app.route('/employee/add', methods=['GET', 'POST'])
@login_required
def add_employee():
    if current_user.admin_level < 1:
        flash('Access denied.')
        return redirect(url_for('home'))
    form = EmployeeForm()
    if form.validate_on_submit():
        emp = Employee(
            first_name=form.first_name.data,
            last_name=form.last_name.data,
            job_title=form.job_title.data,
            line_manager=form.line_manager.data,
            service=form.service.data,
            start_date=form.start_date.data,
            team_id=form.team_id.data,
            email=form.email.data
        )
        db.session.add(emp)
        db.session.commit()
        flash('New employee added.')
        return redirect(url_for('employee_list'))
    return render_template('add_employee.html', form=form)

@app.route('/employee/quick-add', methods=['POST'])
@login_required
@csrf.exempt
def quick_add_employee():
    """
    Minimal quick-add endpoint for use in wizard Step 1.
    Expects JSON: { first_name, last_name, role, service }
    """
    data = request.get_json(force=True, silent=True) or {}
    first = (data.get("first_name") or "").strip()
    last = (data.get("last_name") or "").strip()
    role = (data.get("role") or "").strip()
    service = (data.get("service") or "").strip()

    if not first or not last:
        return jsonify({"success": False, "error": "First and last name are required"}), 400

    emp = Employee(
        first_name=first,
        last_name=last,
        role=role,
        service=service,
        manager_id=getattr(current_user, "id", None)
    )
    db.session.add(emp)
    db.session.commit()

    # Optional: log event
    try:
        evt = TimelineEvent(
            event_type="Employee Created",
            notes="Employee created via Quick-Add in wizard",
            updated_by=current_user.username
        )
        db.session.add(evt)
        db.session.commit()
    except Exception:
        pass

    return jsonify({"success": True, "id": emp.id, "display_name": f"{emp.first_name} {emp.last_name}"})

@app.route('/employee/list')
@login_required
def employee_list():
    # Choose template by active module, but always apply role-based filtering
    template = 'probation_employee_list.html' if session.get('active_module') == 'Probation' else 'employee_list.html'

    q = Employee.query
    if current_user.admin_level == 0:
        # Line managers: only their team
        if current_user.team_id:
            q = q.filter(Employee.team_id == current_user.team_id)
        else:
            # No team assigned -> return empty list rather than error
            q = q.filter(False)

    employees = q.order_by(Employee.last_name.asc(), Employee.first_name.asc()).all()
    return render_template(template, employees=employees)

@app.route('/pip/select-employee', methods=['GET', 'POST'])
@login_required
def select_employee_for_pip():
    employees = Employee.query.order_by(Employee.last_name).all()
    if request.method == 'POST':
        return redirect(url_for('create_pip', employee_id=request.form.get('employee_id')))
    return render_template('pip_select_employee.html', employees=employees)

# ----- Document Generation -----
@app.route('/pip/<int:id>/generate/invite')
@login_required
def generate_invite_letter(id):
    pip = PIPRecord.query.get_or_404(id)
    tpl = DocxTemplate('templates/docx/invite_letter_template.docx')

    context = {
        'employee': pip.employee,
        'pip': pip,
        'current_user': current_user,
        'current_date': datetime.utcnow().strftime('%d %B %Y'),
        'created_by': pip.created_by or current_user.username
    }

    tpl.render(context)
    out = BytesIO()
    tpl.save(out)
    out.seek(0)

    return send_file(
        out,
        as_attachment=True,
        download_name=f"Invite_Letter_{pip.employee.last_name}_{pip.id}.docx",
        mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    )

@app.route('/pip/<int:id>/generate/plan')
@login_required
def generate_plan_document(id):
    pip = PIPRecord.query.get_or_404(id)
    tpl = DocxTemplate('templates/docx/plan_template.docx')
    tpl.render({'employee': pip.employee, 'pip': pip, 'current_user': current_user})
    out = BytesIO()
    tpl.save(out)
    out.seek(0)
    return send_file(
        out,
        as_attachment=True,
        download_name=f"PIP_Plan_{pip.employee.last_name}_{pip.id}.docx",
        mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    )

@app.route('/pip/<int:id>/generate/outcome')
@login_required
def generate_outcome_letter(id):
    pip = PIPRecord.query.get_or_404(id)
    tpl = DocxTemplate('templates/docx/outcome_letter_template.docx')
    date_obj = pip.last_updated or pip.created_at or pip.start_date or datetime.now()
    date_str = date_obj.strftime('%d %b %Y')
    tpl.render({'employee': pip.employee, 'pip': pip, 'current_user': current_user, 'date_str': date_str})
    out = BytesIO()
    tpl.save(out)
    out.seek(0)
    return send_file(
        out,
        as_attachment=True,
        download_name=f"Outcome_Letter_{pip.employee.last_name}_{pip.id}.docx",
        mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    )

# ----- Wizard -----
class DummyForm(FlaskForm):
    pass

@app.route('/pip/create-wizard', methods=['GET', 'POST'])
@login_required
def create_pip_wizard():
    if 'wizard_step' not in session:
        session['wizard_step'] = 1
        session['pip_data'] = {}

    step = session['wizard_step']
    data = session.get('pip_data', {})
    wizard_errors = {}
    draft = None  # Enhance later: load a real draft by draft_id

    # --- Clickable stepper: handle ?goto=N (allow jumps only up to the max allowed) ---
    if request.method == 'GET':
        try:
            goto = int(request.args.get('goto', 0))
        except (TypeError, ValueError):
            goto = 0
        if goto:
            max_allowed = _max_wizard_step(data)
            if 1 <= goto <= max_allowed:
                session['wizard_step'] = goto
                step = goto

    if request.method == 'POST':
        print(f"[DEBUG] POST request received at step {step}")
        print(f"[DEBUG] Form data: {request.form}")

        if step == 1:
            employee_id = request.form.get('employee_id')
            draft_name = request.form.get('draft_name', '').strip()
            if not employee_id:
                wizard_errors['employee_id'] = "Please select an employee."
            else:
                data['employee_id'] = int(employee_id)
            data['draft_name'] = draft_name

        elif step == 2:
            concerns = request.form.get('concerns', '').strip()
            category = request.form.get('concern_category', '').strip()
            severity = request.form.get('severity', '').strip()
            frequency = request.form.get('frequency', '').strip()
            tags = request.form.get('concern_tags', '').strip()
            draft_name = request.form.get('draft_name', '').strip()

            if not concerns:
                wizard_errors['concerns'] = "Concerns cannot be empty."
            if not category:
                wizard_errors['concern_category'] = "Please choose a concern category."
            if not severity:
                wizard_errors['severity'] = "Please select severity."
            if not frequency:
                wizard_errors['frequency'] = "Please select frequency."

            data.update({
                'concerns': concerns,
                'concern_category': category,
                'severity': severity,
                'frequency': frequency,
                'concern_tags': tags,
                'draft_name': draft_name,
            })

        elif step == 3:
            start_date = request.form.get('start_date')
            review_date = request.form.get('review_date')
            draft_name = request.form.get('draft_name', '').strip()

            if not start_date:
                wizard_errors['start_date'] = "Start date is required."
            if not review_date:
                wizard_errors['review_date'] = "Review date is required."

            if not wizard_errors:
                data['start_date'] = start_date
                data['review_date'] = review_date
            data['draft_name'] = draft_name

        elif step == 4:
            data['capability_meeting_date'] = request.form.get('capability_meeting_date')
            data['capability_meeting_time'] = request.form.get('capability_meeting_time')
            data['capability_meeting_venue'] = request.form.get('capability_meeting_venue')
            data['draft_name'] = request.form.get('draft_name', '').strip()

        elif step == 5:
            # Step 5 now ONLY validates and stores the items,
            # the DB commit happens on Step 6 (Review & Submit).
            action_items = request.form.getlist('action_plan_items[]')
            valid_items = [item.strip() for item in action_items if item.strip()]
            if not valid_items:
                wizard_errors['action_plan_items'] = "Add at least one action plan item."
            else:
                data['action_plan_items'] = valid_items
                # Persist and move to Review step
                session['pip_data'] = data
                session['wizard_step'] = 6
                return redirect(url_for('create_pip_wizard'))

        elif step == 6:
            # Final COMMIT — create PIP and Action Items
            try:
                items = data.get('action_plan_items') or []
                if not any((x or '').strip() for x in items):
                    # Safety net: if somehow we got here with no items, send back to step 5
                    flash("Please add at least one action plan item.", "warning")
                    session['wizard_step'] = 5
                    return redirect(url_for('create_pip_wizard'))

                pip = PIPRecord(
                    employee_id=int(data['employee_id']),
                    concerns=data['concerns'],
                    concern_category=data.get('concern_category'),
                    severity=data.get('severity'),
                    frequency=data.get('frequency'),
                    tags=data.get('concern_tags'),
                    start_date=datetime.strptime(data['start_date'], '%Y-%m-%d').date(),
                    review_date=datetime.strptime(data['review_date'], '%Y-%m-%d').date(),
                    capability_meeting_date=datetime.strptime(data['capability_meeting_date'], '%Y-%m-%d')
                        if data.get('capability_meeting_date') else None,
                    capability_meeting_time=data.get('capability_meeting_time'),
                    capability_meeting_venue=data.get('capability_meeting_venue'),
                    created_by=current_user.username
                )
                db.session.add(pip)
                db.session.commit()

                for item_text in items:
                    action = PIPActionItem(pip_record_id=pip.id, description=item_text)
                    db.session.add(action)

                timeline = TimelineEvent(
                    pip_record_id=pip.id,
                    event_type="PIP Created",
                    notes="PIP created via multi-step wizard",
                    updated_by=current_user.username
                )
                db.session.add(timeline)

                db.session.commit()

                session.pop('wizard_step', None)
                session.pop('pip_data', None)

                flash("PIP created successfully!", "success")
                return redirect(url_for('dashboard'))

            except Exception as e:
                print(f"[ERROR] Failed to save PIP: {e}")
                flash(f"Error creating PIP: {str(e)}", "danger")

        # Always keep draft name if present
        if 'draft_name' in request.form:
            data['draft_name'] = request.form.get('draft_name')

        # Persist to session
        session['pip_data'] = data

        # Move to next step if no errors and not at final commit step
        if not wizard_errors and step < 5:
            session['wizard_step'] = step + 1
            return redirect(url_for('create_pip_wizard'))

    employees = Employee.query.all() if step == 1 else []
    print(f"[DEBUG] Rendering step {step} with data: {data}")

    # Compute allowed max step for template (clickable stepper)
    max_allowed = _max_wizard_step(data)

    return render_template(
        'create_pip_wizard.html',
        step=step,
        draft=draft,
        data=data,
        wizard_errors=wizard_errors,
        employees=employees,
        max_allowed_step=max_allowed  # <-- for the clickable stepper UI
    )

# ----- AJAX validation for wizard -----
@app.route('/validate_wizard_step', methods=['POST'])
@login_required
def validate_wizard_step():
    step = int(request.form.get("step", 1))
    errors = {}

    if step == 1:
        employee_id = request.form.get("employee_id")
        if not employee_id:
            errors["employee_id"] = "Please select an employee."

    elif step == 2:
        concerns = request.form.get("concerns", "").strip()
        category = request.form.get("concern_category", "").strip()
        tags = request.form.get("concern_tags", "").strip()
        severity = request.form.get("severity", "").strip()
        frequency = request.form.get("frequency", "").strip()

        if not concerns:
            errors["concerns"] = "Please describe the concern."
        if not category:
            errors["concern_category"] = "Please select a concern category."
        # tags optional to match wizard
        if not severity:
            errors["severity"] = "Please select severity."
        if not frequency:
            errors["frequency"] = "Please select frequency."

    elif step == 3:
        start_date = request.form.get("start_date", "").strip()
        review_date = request.form.get("review_date", "").strip()
        if not start_date:
            errors["start_date"] = "Please enter a start date."
        if not review_date:
            errors["review_date"] = "Please enter a review date."

    elif step == 4:
        meeting_date = request.form.get("meeting_date", "").strip()
        meeting_time = request.form.get("meeting_time", "").strip()
        if not meeting_date:
            errors["meeting_date"] = "Please enter a meeting date."
        if not meeting_time:
            errors["meeting_time"] = "Please enter a meeting time."

    elif step == 5:
        actions = request.form.getlist("action_plan_items[]")
        actions = [a.strip() for a in actions if a.strip()]
        if not actions:
            errors["action_plan_items"] = "Please add at least one action item."

    # Step 6 has no extra client-side fields; server handles final commit

    return jsonify({"success": not errors, "errors": errors})

# ----- Concern taxonomy (categories + tag suggestions) -----
@app.route('/taxonomy/categories', methods=['GET'])
@login_required
def taxonomy_categories():
    """
    Simple static set to start with. You can later load from DB/Settings.
    """
    categories = [
        "Timekeeping",
        "Attendance",
        "Quality of Work",
        "Productivity",
        "Conduct",
        "Communication",
        "Teamwork/Collaboration",
        "Compliance/Process",
        "Customer Service",
        "Health & Safety"
    ]
    return jsonify({"categories": categories})

@app.route('/taxonomy/tags_suggest', methods=['GET'])
@login_required
def taxonomy_tags_suggest():
    """
    Suggest tags from curated set (by ?category=) + recent PIP tags (comma lists), optional filter ?q=
    """
    q = (request.args.get('q') or '').strip().lower()
    category = (request.args.get('category') or '').strip()

    # Recent from DB
    try:
        recent_rows = db.session.query(PIPRecord.tags).order_by(PIPRecord.id.desc()).limit(200).all()
    except Exception:
        recent_rows = []

    recent = []
    for (tag_str,) in recent_rows:
        if not tag_str: continue
        for t in (tag_str.split(',') if isinstance(tag_str, str) else []):
            t = (t or '').strip()
            if t: recent.append(t)

    merged = _merge_curated_and_recent(category, recent, cap=40)

    if q:
        merged = [t for t in merged if q in t.lower()]

    return jsonify({"tags": merged[:30]})

# ----- Draft handling -----
@app.route('/dismiss_draft', methods=['POST'])
@login_required
@csrf.exempt
def dismiss_draft():
    draft = DraftPIP.query.filter_by(user_id=current_user.id, is_dismissed=False).first()
    if draft:
        draft.is_dismissed = True
        draft.updated_at = datetime.now(timezone.utc)
        db.session.commit()
        return jsonify({"success": True})
    return jsonify({"success": False, "message": "No active draft found"}), 404

@app.route('/save_pip_draft', methods=['POST'])
@login_required
@csrf.exempt
def save_pip_draft():
    try:
        data = {
            'employee_id': request.form.get('employee_id'),
            'draft_name': request.form.get('draft_name', '').strip(),
            'concerns': request.form.get('concerns', '').strip(),
            'concern_category': request.form.get('concern_category', '').strip(),
            'severity': request.form.get('severity', '').strip(),
            'frequency': request.form.get('frequency', '').strip(),
            'concern_tags': request.form.get('concern_tags', '').strip(),
            'start_date': request.form.get('start_date'),
            'review_date': request.form.get('review_date'),
            'capability_meeting_date': request.form.get('capability_meeting_date'),
            'capability_meeting_time': request.form.get('capability_meeting_time'),
            'capability_meeting_venue': request.form.get('capability_meeting_venue'),
            'action_plan_items': request.form.getlist('action_plan_items[]')
        }

        cleaned_data = {k: v for k, v in data.items() if v not in [None, '', []]}

        existing_draft = DraftPIP.query.filter_by(user_id=current_user.id, is_dismissed=False).first()
        if existing_draft:
            db.session.delete(existing_draft)
            db.session.commit()

        new_draft = DraftPIP(
            user_id=current_user.id,
            data=cleaned_data,
            step=session.get('wizard_step', 1),
            is_dismissed=False,
            updated_at=datetime.now(timezone.utc)
        )
        db.session.add(new_draft)
        db.session.commit()

        return jsonify({"success": True, "message": "Draft saved."})

    except Exception as e:
        print(f"[ERROR] Failed to save draft: {e}")
        return jsonify({"success": False, "message": "Failed to save draft."}), 500

# ----- AI Action Suggestions -----
@app.route('/suggest_actions_ai', methods=['POST'])
@login_required
def suggest_actions_ai():
    data = request.get_json() or {}
    concerns = (data.get('concerns') or '').strip()
    severity = (data.get('severity') or '').strip()
    frequency = (data.get('frequency') or '').strip()
    tags = (data.get('tags') or '').strip()
    category = (data.get('category') or '').strip()

    def _dedupe_clean(items):
        out, seen = [], set()
        for x in (items or []):
            s = (x or '').strip()
            if not s:
                continue
            key = s.lower()
            if key not in seen:
                out.append(s)
                seen.add(key)
        return out[:8]  # keep it tidy

    # ----- Build a strict-JSON prompt -----
    sys_msg = (
        "You are an HR advisor in the UK. "
        "Return ONLY valid JSON with two arrays: "
        '{"actions": ["short concrete manager actions"], "next_up": ["quick follow-ups or escalations"]}. '
        "Actions must be specific, measurable where possible, and supportive. "
        "No prose, no markdown, JSON only."
    )
    user_msg = f"""
Concern Category: {category or "[unspecified]"}
Concerns: {concerns or "[none]"}
Tags: {tags or "[none]"}
Severity: {severity or "[unspecified]"}
Frequency: {frequency or "[unspecified]"}

Rules:
- Provide 3–5 'actions' tailored to the inputs.
- Provide 2–4 'next_up' items (e.g., monitoring cadence, policy references, escalation steps).
- Keep each item under 140 characters.
- JSON ONLY.
"""

    actions, next_up = [], []

    try:
        resp = client.chat.completions.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": sys_msg},
                {"role": "user", "content": user_msg}
            ],
            temperature=0.5,
            max_tokens=300
        )
        raw = (resp.choices[0].message.content or "").strip()

        # Try strict JSON parse
        import json, re
        # Extract the first {...} block just in case
        m = re.search(r"\{[\s\S]*\}", raw)
        payload = json.loads(m.group(0) if m else raw)

        actions = _dedupe_clean(payload.get("actions", []))
        next_up = _dedupe_clean(payload.get("next_up", []))

    except Exception as e:
        # Fallback: simple heuristic parsing if model didn't return JSON
        # Split lines, grab up to 5 actions
        lines = [ln.strip("-•* 0123456789.\t") for ln in (raw.splitlines() if 'raw' in locals() else [])]
        actions = _dedupe_clean([ln for ln in lines if ln][:5])

    # Server-side "Next Up" enrichment (defensive, additive)
    tag_list = [t.strip().lower() for t in tags.split(",")] if tags else []
    cat = (category or "").lower()
    sev = (severity or "").lower()
    freq = (frequency or "").lower()

    suggestions = []
    if 'lateness' in tag_list or 'timekeeping' in cat:
        suggestions += [
            "Daily start-time check-ins for 2 weeks",
            "Agree punctuality targets; log variances",
        ]
    if 'conduct' in tag_list or cat == 'conduct':
        suggestions += [
            "Reference conduct policy; document conversations",
            "Book values/behaviour refresher"
        ]
    if 'performance' in cat or 'missed deadlines' in tags.lower():
        suggestions += [
            "Weekly milestones with due dates",
            "Stand-up updates Mon/Wed/Fri"
        ]
    if sev == 'high':
        suggestions += ["Escalate to formal stage if no progress"]
    if freq in ('frequent', 'persistent'):
        suggestions += ["Increase monitoring and assign a buddy/mentor"]

    # Merge + clean
    next_up = _dedupe_clean((next_up or []) + suggestions)

    return jsonify({
        "success": True,
        "actions": actions,
        "next_up": next_up
    }), 200

# ----- Probation module -----
@app.route('/probation/<int:id>')
@login_required
def view_probation(id):
    probation = ProbationRecord.query.get_or_404(id)
    employee = probation.employee
    return render_template('view_probation.html', probation=probation, employee=employee)

@app.route('/probation/<int:id>/review/add', methods=['GET', 'POST'])
@login_required
def add_probation_review(id):
    probation = ProbationRecord.query.get_or_404(id)
    form = ProbationReviewForm()

    if form.validate_on_submit():
        review = ProbationReview(
            probation_id=probation.id,
            review_date=form.review_date.data,
            reviewer=form.reviewer.data,
            summary=form.summary.data,
            concerns_flag=(form.concerns_flag.data.lower() == 'yes')
        )
        db.session.add(review)

        event = TimelineEvent(
            pip_record_id=None,
            event_type="Probation Review",
            notes=f"Review added by {current_user.username}",
            updated_by=current_user.username
        )
        db.session.add(event)

        db.session.commit()
        flash('Probation review added.', 'success')
        return redirect(url_for('view_probation', id=probation.id))

    return render_template('add_probation_review.html', form=form, probation=probation)

@app.route('/probation/<int:id>/plan/add', methods=['GET', 'POST'])
@login_required
def add_probation_plan(id):
    probation = ProbationRecord.query.get_or_404(id)
    form = ProbationPlanForm()

    if form.validate_on_submit():
        plan = ProbationPlan(
            probation_id=probation.id,
            objectives=form.objectives.data,
            deadline=form.deadline.data,
            outcome=form.outcome.data
        )
        db.session.add(plan)

        event = TimelineEvent(
            pip_record_id=None,
            event_type="Probation Plan Added",
            notes=f"Plan created by {current_user.username}",
            updated_by=current_user.username
        )
        db.session.add(event)

        db.session.commit()
        flash('Development plan added.', 'success')
        return redirect(url_for('view_probation', id=probation.id))

    return render_template('add_probation_plan.html', form=form, probation=probation)

@app.route('/probation/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_probation(id):
    probation = ProbationRecord.query.get_or_404(id)
    form = ProbationRecordForm(obj=probation)

    if form.validate_on_submit():
        probation.start_date = form.start_date.data
        probation.expected_end_date = form.expected_end_date.data
        probation.notes = form.notes.data
        db.session.commit()
        flash('Probation record updated.', 'success')
        return redirect(url_for('view_probation', id=probation.id))

    return render_template('edit_probation.html', form=form, probation=probation)

@app.route('/probation/<int:id>/status/<new_status>', methods=['POST'])
@login_required
def update_probation_status(id, new_status):
    probation = ProbationRecord.query.get_or_404(id)
    valid_statuses = ['Completed', 'Extended', 'Failed']

    if new_status not in valid_statuses:
        flash('Invalid status update.', 'danger')
        return redirect(url_for('view_probation', id=id))

    probation.status = new_status
    db.session.add(probation)

    event = TimelineEvent(
        pip_record_id=None,
        event_type="Probation Status Updated",
        notes=f"Status changed to {new_status} by {current_user.username}",
        updated_by=current_user.username
    )
    db.session.add(event)

    db.session.commit()
    flash(f'Status updated to {new_status}.', 'success')
    return redirect(url_for('view_probation', id=id))

@app.route('/probation/create/<int:employee_id>', methods=['GET', 'POST'])
@login_required
def create_probation(employee_id):
    employee = Employee.query.get_or_404(employee_id)
    form = ProbationRecordForm()

    if form.validate_on_submit():
        probation = ProbationRecord(
            employee_id=employee.id,
            start_date=form.start_date.data,
            expected_end_date=form.expected_end_date.data,
            notes=form.notes.data
        )
        db.session.add(probation)
        db.session.commit()
        flash('Probation record created successfully.', 'success')
        return redirect(url_for('view_probation', id=probation.id))

    return render_template('create_probation.html', form=form, employee=employee)

@app.route('/probation/dashboard')
@login_required
def probation_dashboard():
    records = ProbationRecord.query.all()
    return render_template('probation_dashboard.html', records=records)

@app.route('/probation/employees')
@login_required
def probation_employee_list():
    session['active_module'] = 'Probation'
    employees = Employee.query.all()
    return render_template('probation_employee_list.html', employees=employees)

@app.route('/admin/users')
@login_required
def manage_users():
    if not current_user.is_superuser():
        flash('Access denied.', 'danger')
        return redirect(url_for('dashboard'))
    users = User.query.all()
    return render_template('admin_users.html', users=users)

@app.route('/admin/backup')
@login_required
def backup_database():
    if not current_user.is_superuser():
        flash('Access denied: Superuser only.', 'danger')
        return redirect(url_for('home'))

    db_path = os.path.join(os.getcwd(), 'pip_crm.db')
    if os.path.exists(db_path):
        return send_file(db_path, as_attachment=True)
    else:
        flash('Database file not found.', 'danger')
        return redirect(url_for('admin_dashboard'))

@app.route("/employee/import", methods=["GET", "POST"])
@login_required
@superuser_required
def employee_import():
    """
    GET: render basic upload page (you can wire a simple template later).
    POST: read file, return preview (headers + first 10 rows), and a temp_id.
    """
    if request.method == "GET":
        # If you have a template, render it; otherwise return a tiny placeholder.
        return render_template("employee_import.html")  # create later

    file = request.files.get("file")
    if not file or file.filename == "":
        abort(400, "No file uploaded")

    filename = secure_filename(file.filename)
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext not in ALLOWED_EXTS:
        abort(400, f"Unsupported file type: .{ext}")

    file_bytes = file.read()

    if ext == "csv":
        headers, rows = _read_csv_bytes(file_bytes)
    else:
        headers, rows = _read_xlsx_bytes(file_bytes)

    # Basic normalised headers for mapping convenience
    normalised_headers = [{ "raw": h, "norm": _normalize_header(h) } for h in (headers or [])]

    # Stash raw file in a temp file; return a token the browser will send back later
    tmp = tempfile.NamedTemporaryFile(prefix="emp_import_", suffix=f".{ext}", delete=False)
    tmp.write(file_bytes)
    tmp.flush()
    temp_id = tmp.name  # path as token; you can swap to DB-backed jobs later

    # Only preview first 10 rows to keep payload small
    preview = rows[:10] if rows else []

    return jsonify({
        "temp_id": temp_id,
        "headers": headers,
        "headers_norm": normalised_headers,
        "suggested_mapping": _suggest_mapping(headers),
        "preview_rows": preview,
        "xlsx_enabled": XLSX_ENABLED
    }), 200

def _suggest_mapping(headers):
    mapping = {}
    for h in headers or []:
        n = _normalize_header(h)
        if n in EMPLOYEE_FIELDS:
            mapping[h] = n
        elif n in ("firstname", "first"):
            mapping[h] = "first_name"
        elif n in ("lastname", "last", "surname"):
            mapping[h] = "last_name"
        elif n in ("mail", "email_address"):
            mapping[h] = "email"
        elif n in ("role", "position", "title", "job", "jobrole", "job_role", "jobtitle"):
            mapping[h] = "job_title"
        elif n in ("manager", "line_manager", "linemanager", "manager_name"):
            mapping[h] = "line_manager"
        elif n in ("team", "teamid", "team_id", "dept", "department"):
            mapping[h] = "team_id"
        else:
            mapping[h] = ""  # not mapped yet
    return mapping

@app.route("/employee/import/validate", methods=["POST"])
@login_required
@superuser_required
def employee_import_validate():
    """
    Body JSON:
    {
      "temp_id": "<tmp path from /employee/import>",
      "mapping": { "<file header>": "<employee_field>" },
      "unique_key": "email" | "first_name,last_name"
    }
    Returns a validation report (missing requireds, duplicates, unmapped).
    """
    data = request.get_json(force=True, silent=True) or {}
    temp_id = data.get("temp_id")
    mapping = data.get("mapping") or {}
    unique_key = (data.get("unique_key") or "email").strip()

    if not temp_id:
        abort(400, "Missing temp_id")
    try:
        with open(temp_id, "rb") as f:
            file_bytes = f.read()
    except Exception:
        abort(400, "Invalid temp_id or temporary file expired")

    ext = temp_id.rsplit(".", 1)[-1].lower()
    headers, rows = (_read_csv_bytes(file_bytes) if ext == "csv" else _read_xlsx_bytes(file_bytes))

    # Build normalised row dicts by applying mapping
    mapped_rows = []
    unmapped_headers = []
    for h in headers or []:
        if not mapping.get(h):
            unmapped_headers.append(h)

    for r in rows:
        out = {}
        for h, v in r.items():
            field = mapping.get(h)
            if not field:
                continue
            if field == "start_date":
                out[field] = _try_parse_date(v)
            else:
                out[field] = (str(v).strip() if v is not None else None)
        mapped_rows.append(out)

    # Required fields check
    missing_required = []
    for idx, r in enumerate(mapped_rows, start=1):
        missing = [f for f in REQUIRED_FIELDS if not r.get(f)]
        if missing:
            missing_required.append({"row": idx, "missing": missing})

    # Duplicate checks (in-file)
    duplicates_in_file = []
    if unique_key:
        keys = unique_key.split(",")
        seen = set()
        for idx, r in enumerate(mapped_rows, start=1):
            key_tuple = tuple((r.get(k.strip()) or "").lower() for k in keys)
            if all(key_tuple):  # only consider non-empty key
                if key_tuple in seen:
                    duplicates_in_file.append({"row": idx, "key": key_tuple})
                else:
                    seen.add(key_tuple)

    # Duplicate checks (in DB) — only if key fields map to actual columns we have
    duplicates_in_db = []
    try:
        from sqlalchemy import or_, and_
        # Build a query for emails or name pair
        if unique_key == "email" and any(r.get("email") for r in mapped_rows):
            emails = list({r.get("email") for r in mapped_rows if r.get("email")})
            existing = set(e[0].lower() for e in db.session.query(Employee.email).filter(Employee.email.in_(emails)).all())
            for idx, r in enumerate(mapped_rows, start=1):
                em = (r.get("email") or "").lower()
                if em and em in existing:
                    duplicates_in_db.append({"row": idx, "email": r.get("email")})
        elif unique_key == "first_name,last_name":
            # Crude name-based check
            pairs = {( (r.get("first_name") or "").lower(), (r.get("last_name") or "").lower() )
                     for r in mapped_rows if r.get("first_name") and r.get("last_name")}
            if pairs:
                q = db.session.query(Employee.first_name, Employee.last_name).all()
                existing_pairs = {(fn.lower(), ln.lower()) for fn, ln in q}
                for idx, r in enumerate(mapped_rows, start=1):
                    t = ((r.get("first_name") or "").lower(), (r.get("last_name") or "").lower())
                    if all(t) and t in existing_pairs:
                        duplicates_in_db.append({"row": idx, "name": f"{r.get('first_name')} {r.get('last_name')}"})
    except Exception as e:
        # Don’t fail validation if the DB check had an issue (e.g., missing column). Report it.
        duplicates_in_db = [{"error": f"DB duplicate check skipped: {e}"}]

    report = {
        "unmapped_headers": unmapped_headers,
        "missing_required": missing_required,
        "duplicates_in_file": duplicates_in_file,
        "duplicates_in_db": duplicates_in_db,
        "rows_ready": len(mapped_rows) - len(missing_required) - len(duplicates_in_file) - len(duplicates_in_db),
        "total_rows": len(mapped_rows)
    }
    # Keep the mapped rows for the commit step (send them back to client, or if preferred store server-side)
    return jsonify({"temp_id": temp_id, "report": report}), 200

@app.route("/employee/import/commit", methods=["POST"])
@login_required
@superuser_required
def employee_import_commit():
    """
    Body JSON:
    {
      "temp_id": "<tmp path>",
      "mapping": { "<file header>": "<employee_field>" },
      "unique_key": "email" | "first_name,last_name",
      "confirm": true
    }
    Creates Employee rows for all *valid* entries.
    """
    data = request.get_json(force=True, silent=True) or {}
    temp_id = data.get("temp_id")
    mapping = data.get("mapping") or {}
    confirm = bool(data.get("confirm"))
    unique_key = (data.get("unique_key") or "email").strip()
    if not (temp_id and confirm and mapping):
        abort(400, "Missing temp_id, mapping, or confirm flag")

    try:
        with open(temp_id, "rb") as f:
            file_bytes = f.read()
    except Exception:
        abort(400, "Invalid temp_id or temporary file expired")

    ext = temp_id.rsplit(".", 1)[-1].lower()
    headers, rows = (_read_csv_bytes(file_bytes) if ext == "csv" else _read_xlsx_bytes(file_bytes))

    created, skipped, errors = 0, 0, []

    for idx, src in enumerate(rows, start=1):
        payload = {}
        for h, v in src.items():
            field = mapping.get(h)
            if not field:
                continue
            if field == "start_date":
                payload[field] = _try_parse_date(v)
            else:
                payload[field] = (str(v).strip() if v is not None else None)

        # Required fields
        if any(not payload.get(f) for f in REQUIRED_FIELDS):
            skipped += 1
            continue

        # Duplicate checks (quick)
        try:
            exists = False
            if unique_key == "email" and payload.get("email"):
                exists = db.session.query(Employee.id).filter_by(email=payload["email"]).first() is not None
            elif unique_key == "first_name,last_name" and payload.get("first_name") and payload.get("last_name"):
                exists = db.session.query(Employee.id).filter_by(
                    first_name=payload["first_name"], last_name=payload["last_name"]
                ).first() is not None
            if exists:
                skipped += 1
                continue
        except Exception as e:
            errors.append({"row": idx, "error": f"Duplicate check failed: {e}"})
            skipped += 1
            continue

        try:
            emp = Employee(**{k: v for k, v in payload.items() if k in EMPLOYEE_FIELDS})
            db.session.add(emp)
            created += 1
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})
            skipped += 1

    db.session.commit()

    # Log a single summary event (adjust if you prefer per-employee)
    try:
        # Replace with your current_user import if needed
        username = getattr(current_user, "username", "system")
        notes = f"Employee Import: created={created}, skipped={skipped}, errors={len(errors)}"
        evt = TimelineEvent(event_type="Import", notes=notes, updated_by=username)
        db.session.add(evt)
        db.session.commit()
    except Exception:
        pass  # If TimelineEvent doesn’t fit here, skip logging silently

    return jsonify({"created": created, "skipped": skipped, "errors": errors}), 200

def _parse_numbered_list(text, max_items=6):
        """
        Turn LLM output into a clean list of items.
        Accepts '1. Do X', '- Do X', or plain lines.
        """
        if not text:
            return []
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        items = []
        for l in lines:
            # Strip leading bullets / numbering like "1) ", "2. ", "- ", "• "
            l = re.sub(r'^\s*(?:[-*•]|\d+[\).\]]?)\s*', '', l).strip()
            if l:
                items.append(l)
            if len(items) >= max_items:
                break
        # Deduplicate while preserving order
        seen = set()
        out = []
        for it in items:
            if it not in seen:
                out.append(it)
                seen.add(it)
        return out


@app.route('/pip/wizard/resume', methods=['GET'])
@login_required
def pip_wizard_resume():
    """Hydrate the wizard session from the latest active DraftPIP and jump to the furthest valid step."""
    draft = get_active_draft_for_user(current_user.id)
    if not draft or not draft.data:
        flash("No active draft to resume.", "warning")
        return redirect(url_for('dashboard'))

    # Load draft into session and jump to the max allowed step
    session['pip_data'] = dict(draft.data)
    session['wizard_step'] = _max_wizard_step(session['pip_data'])
    return redirect(url_for('create_pip_wizard'))

# Create DB if missing
with app.app_context():
    if not os.path.exists(os.path.join(basedir, 'pip_crm.db')):
        db.create_all()
        print('✅ Database created')

@app.route('/ping')
def ping():
    return 'Pong!'

print('✅ Flask app initialized and ready.')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
