from __future__ import annotations

import csv
import io
from datetime import datetime

try:
    import openpyxl  # noqa: F401
    XLSX_ENABLED = True
except Exception:
    openpyxl = None
    XLSX_ENABLED = False

ALLOWED_EXTS = {"csv", "xlsx"} if XLSX_ENABLED else {"csv"}

EMPLOYEE_FIELDS = [
    "first_name", "last_name", "email", "job_title", "line_manager",
    "service", "team_id", "start_date"
]
REQUIRED_FIELDS = ["first_name", "last_name"]


def read_csv_bytes(file_bytes: bytes):
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = [dict(r) for r in reader]
    headers = reader.fieldnames or []
    return headers, rows


def read_xlsx_bytes(file_bytes: bytes):
    if not XLSX_ENABLED or openpyxl is None:
        raise RuntimeError("XLSX support is not available")

    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes))
    sheet = workbook.active
    headers = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append({headers[i]: (row[i] if i < len(headers) else None) for i in range(len(headers))})
    return headers, rows


def normalize_header(h):
    return (h or "").strip().lower().replace(" ", "_")


def try_parse_date(value):
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except Exception:
            continue
    return None


def parse_iso_date(s):
    try:
        return datetime.strptime((s or "").strip(), "%Y-%m-%d").date()
    except Exception:
        return None